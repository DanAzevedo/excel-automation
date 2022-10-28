from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl import Workbook

font = Font(name='Calibri',
            size=12,
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='FF000000')

fill = PatternFill(fill_type=None,
                   start_color='FFFFFFFF',
                   end_color='FF000000')

'''
Usar um entre:
{'thin', 'dashed', 'mediumDashDot', 'dashDotDot', 'hair',
'dotted', 'mediumDashDotDot', 'medium', 'double', 'dashDot',
'slantDashDot', 'thick', 'mediumDashed'} 
'''
border = Border(left=Side(border_style=None,
                          color='FF000000'),
                right=Side(border_style='double',
                           color='FF0000FF'),
                top=Side(border_style=None,
                         color='FF000000'),
                bottom=Side(border_style=None,
                            color='FF000000'),
                diagonal=Side(border_style=None,
                              color='FF000000'),
                diagonal_direction=0,
                outline=Side(border_style=None,
                             color='FF000000'),
                vertical=Side(border_style=None,
                              color='FF000000'),
                horizontal=Side(border_style=None,
                                color='FF000000')
                )

'''
{'justify', 'top', 'center', 'distributed', 'bottom'}
{'distributed', 'top', 'justify', 'center', 'bottom'}
'''
alignment = Alignment(horizontal='center',
                      vertical='bottom',
                      text_rotation=0,
                      wrap_text=False,
                      shrink_to_fit=False,
                      indent=0)

'''
Data, Decimal, Moeda...
'''
number_format = 'General'

protection = Protection(locked=True,
                        hidden=False)

wb = Workbook()

sheet = wb.active

sheet['A1'] = 'test'
sheet['A2'] = 'test2'
sheet['A3'] = 'test3'

sheet['A1'].font = Font(bold=True, size=20)
sheet['A2'].font = Font(italic=True, color='FFFF0000')  # RGB

sheet['A1'].fill = fill

sheet['A2'].border = border

sheet['A3'].alignment = alignment

wb.save('formatacao.xlsx')
